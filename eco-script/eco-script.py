from openpyxl.styles import PatternFill
from openpyxl import load_workbook


def get_element_summ(book, last_row, element):
	sheet = book.active
	previous_izav = sheet['A2'].value
	summ_kg_year = 0.0
	summ_gr_sec = 0.0
	summ_data_dict = {}

	for row in sheet.iter_rows(min_row = 2, max_col = 6, max_row = last_row, values_only = True):
		contain_element = element in row[3]
		curr_izav = previous_izav if row[0] == None else row[0]
		
		if curr_izav == previous_izav:
			if contain_element:
				summ_kg_year = summ_kg_year + row[4]
				summ_gr_sec = summ_gr_sec + row[5]
		else:
			if summ_kg_year != 0.0 or summ_gr_sec != 0.0:
				summ_data_dict[previous_izav] = [element, summ_kg_year, summ_gr_sec]
			if contain_element:
				summ_kg_year = row[4]
				summ_gr_sec = row[5]
			else:
				summ_kg_year = 0.0
				summ_gr_sec = 0.0

		previous_izav = curr_izav
	return summ_data_dict


def compare_total_book_with_init_dict(book, last_row, summ_data_dict, element):
	sheet = book.active
	previous_izav = sheet['A2'].value

	for row in sheet.iter_rows(min_row = 2, max_col = 8, max_row = last_row):
		contain_element = element in row[3].value
		curr_izav = previous_izav if row[0].value == None else row[0].value
		
		if contain_element and curr_izav in summ_data_dict.keys():
			row[6].value = summ_data_dict[curr_izav][1]
			row[7].value = summ_data_dict[curr_izav][2]
			if row[4].value != row[6].value:
				row[6].fill = PatternFill("solid", fgColor="00FF8080")
			else:
				row[6].fill = PatternFill("solid", fgColor="00CCFFCC")
			if row[5].value != row[7].value:
				row[7].fill = PatternFill("solid", fgColor="00FF8080")
			else:
				row[7].fill = PatternFill("solid", fgColor="00CCFFCC")

		previous_izav = curr_izav

	return book
	

def main():
	element = 'азот'
	input_file_path_1 = 'init-eco.xlsx'
	input_file_path_2 = 'total-eco.xlsx'
	output_file_path = 'comparison.xlsx'

	book_1 = load_workbook(input_file_path_1)
	book_2 = load_workbook(input_file_path_2)

	LAST_ROW_1 = 97
	LAST_ROW_2 = 89

	summ_data_dict = get_element_summ(book_1, LAST_ROW_1, element)
	updated_book = compare_total_book_with_init_dict(book_2, LAST_ROW_2, summ_data_dict, element)

	updated_book.save(output_file_path)
	print('Comparison has been successfully saved in the file "comparison.xlsx".')


if __name__ == "__main__":
	main()
